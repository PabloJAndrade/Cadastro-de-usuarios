from os import name
from sqlite3 import Row
from tkinter.tix import COLUMN
from turtle import title, width
import customtkinter as ctk
import tkinter
from tkinter import END, Frame, StringVar, messagebox
import openpyxl
import xlrd
import pathlib
from openpyxl import Workbook

# Configurando o style padrão do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearance()
        self.all_system()

    def layout_config(self):
        self.title("Sistema de Cadastro de Clientes")
        self.geometry("700x500")

    def appearance(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', '#fff'])
        self.lb_apm.place(x=50, y=430)

        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_appearance)
        self.opt_apm.place(x=50, y=460)

    def all_system(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal")
        frame.place(x=0, y=10)

        title_label = ctk.CTkLabel(frame, text="Sistema de Cadastro de Clientes", font=('Century Gothic bold', 24), text_color="#fff")
        title_label.place(x=190, y=10)

        span = ctk.CTkLabel(self, text="Por favor preencha todos os campos do formulário", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        span.place(x=50, y=70)

        ficheiro = pathlib.Path("Clientes.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro = Workbook()
            folha = ficheiro.active
            folha['A1'] = "Nome Completo"   # type: ignore
            folha['B1'] = "Contato" # type: ignore
            folha['C1'] = "Idade"   # type: ignore
            folha['D1'] = "Genêro"  # type: ignore
            folha['E1'] = "Observações" # type: ignore
            folha['F1'] = "Email"   # type: ignore
            folha['G1'] = "Endereço" # type: ignore

            ficheiro.save("Clientes.xlsx") # type: ignore

        def submit():
            # Pegando os dados das entrys
            name1 = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            gender = gender_combobox.get()
            obs = obs_entry.get(1.0, END)
            email = email_value.get()
            address = address_value.get()

            ficheiro = openpyxl.load_workbook('Clientes.xlsx')
            folha = ficheiro.active
            folha.cell(column=1, row=folha.max_row+1, value=name1)
            folha.cell(column=2, row=folha.max_row, value=contact)
            folha.cell(column=3, row=folha.max_row, value=age)
            folha.cell(column=4, row=folha.max_row, value=gender)
            folha.cell(column=5, row=folha.max_row, value=obs)
            folha.cell(column=6, row=folha.max_row, value=email)
            folha.cell(column=7, row=folha.max_row, value=address)
            
            ficheiro.save(r"Clientes.xlsx")
            messagebox.showinfo("Systema", "Dados salvos com sucesso")
        def clear():
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            obs_entry.delete(1.0, END)
            email_value.set("")
            address_value.set("")

        # variaveis de texto
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()
        email_value = StringVar()

        # Entrys
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic bold", 15), fg_color="transparent")
        contact_entry = ctk.CTkEntry(self, width=200, textvariable=contact_value, font=("Century Gothic bold", 15), fg_color="transparent")
        address_entry = ctk.CTkEntry(self, width=200, textvariable=address_value, font=("Century Gothic bold", 15), fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width=150, textvariable=age_value, font=("Century Gothic bold", 15), fg_color="transparent")

        # Combobox para seletor de genero
        gender_combobox = ctk.CTkComboBox(self, values=['Adão', 'Eva'], font=("Century Gothic bold", 14), width=150)
        gender_combobox.set("Adão")
        email_entry = ctk.CTkEntry(self, width=350, textvariable=email_value, font=("Century Gothic bold", 15), fg_color="transparent")

        # Entrada de Observação
        obs_entry = ctk.CTkTextbox(self, width=470, height=100, font=("arial", 18), border_color="#aaa", border_width=2,
                                   fg_color="transparent")

        # Labels
        lb_name = ctk.CTkLabel(self, text="Digite o seu nome completo", font=("Century Gothic bold", 16),
                               text_color=["#000", "#fff"])
        lb_contact = ctk.CTkLabel(self, text="Digite seu telefone", font=("Century Gothic bold", 16),
                                  text_color=["#000", "#fff"])
        lb_age = ctk.CTkLabel(self, text="Informe sua idade", font=("Century Gothic bold", 16),
                              text_color=["#000", "#fff"])
        lb_gender = ctk.CTkLabel(self, text="Gênero", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_address = ctk.CTkLabel(self, text="Por favor digite seu endereço", font=("Century Gothic bold", 16),
                                  text_color=["#000", "#fff"])
        lb_observation = ctk.CTkLabel(self, text="Observação", font=("Century Gothic bold", 16),
                                      text_color=["#000", "#fff"])
        lb_email = ctk.CTkLabel(self, text="Digite seu melhor e-mail", font=("Century Gothic bold", 16),
                                text_color=["#000", "#fff"])

        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131")
        btn_submit.place(x=320, y=460)
        btn_clear = ctk.CTkButton(self, text="Limpar Campos".upper(), command=clear, fg_color="#555", hover_color="#333")
        btn_clear.place(x=500, y=460)

        # Posição dos elementos na janela
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contact.place(x=450, y=120)
        contact_entry.place(x=450, y=150)

        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)

        lb_gender.place(x=500, y=190)
        gender_combobox.place(x=500, y=220)

        lb_address.place(x=50, y=190)
        address_entry.place(x=50, y=220)

        lb_email.place(x=50, y=260)
        email_entry.place(x=50, y=290)

        lb_observation.place(x=50, y=350)
        obs_entry.place(x=150, y=350)

    def change_appearance(self, new_appearance_mode):
        ctk.set_appearance_mode(new_appearance_mode)

if __name__ == "__main__":
    app = App()
    app.mainloop()
